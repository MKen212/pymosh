# Import Excel Packages
import openpyxl as xl

# Get specific chart Classes separately
from openpyxl.chart import BarChart, Reference

# Process Workbook function
def process_workbook(filename):
    # Load Workbook
    full_filename = filename + ".xlsx"
    wb = xl.load_workbook(full_filename)

    # Specify Sheet to use
    sheet = wb["Sheet1"]
    
    # Loop through rows/cells & fix values
    for row in range(2, sheet.max_row + 1):  # Start from row 2 as row 1 is header, and INCLUDE last row using max + 1
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    # Fix Header
    header_cell = sheet.cell(1, 4)
    header_cell.value = "corrected_price"

    # Select the data to populate the chart
    values = Reference(sheet,
                    min_row=2,
                    max_row=sheet.max_row,
                    min_col=4,
                    max_col=4)

    # Create the chart
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "F2")

    # Save updated workbook & return
    new_filename = filename + "_new.xlsx"
    wb.save(new_filename)
    return new_filename


# Import Path class
from pathlib import Path

# List xlsx files in current directory
path = Path()
for file in path.glob("*.xlsx"):
    print(file)

# Input filename
file_chosen = str(input("Enter Filename to convert (without .xlsx extension)> "))

new_filename = process_workbook(file_chosen)

print(f"Process completed. See: {new_filename}")
